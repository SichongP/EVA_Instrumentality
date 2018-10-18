#!/bin/usr/python2
import openpyxl
from .project import *

def write_opt_att(row_num, key, dict, tab):
	if key == 'PUBLICATION':
		tab.cell(row=row_num, column=6, value=dict[key])
	elif key == 'PARENT':
		tab.cell(row=row_num, column=7, value=dict[key])
	elif key == 'CHILD':
		tab.cell(row=row_num, column=8, value=dict[key])
	elif key == 'PEER':
		tab.cell(row=row_num, column=9, value=dict[key])
	elif key == 'PRJ_LINK':
		tab.cell(row=row_num, column=10, value=dict[key])
	elif key == 'HOLD_DATE':
		tab.cell(row=row_num, column=11, value=dict[key])
	elif key == 'COLLABORATORS':
		tab.cell(row=row_num, column=12, value=dict[key])
	elif key == 'STRAIN':
		tab.cell(row=row_num, column=13, value=dict[key])
	elif key == 'BREED':
		tab.cell(row=row_num, column=14, value=dict[key])
	elif key == 'BROKER':
		tab.cell(row=row_num, column=15, value=dict[key])
#'PLATFORM', 'SOFTWARE', 'PIPELINE', 'IMPUTATION', 'PHASING', 'CENTER', 'DATE', 'ANL_LINK', 'RUN',
	elif key == 'PLATFORM':
		tab.cell(row=row_num, column=8, value=dict[key])
	elif key == 'SOFTWARE':
		tab.cell(row=row_num, column=9, value=dict[key])
	elif key == 'PIPELINE':
		tab.cell(row=row_num, column=10, value=dict[key])
	elif key == 'IMPUTATION':
		tab.cell(row=row_num, column=11, value=dict[key])
	elif key == 'PHASING':
		tab.cell(row=row_num, column=12, value=dict[key])
	elif key == 'CENTER':
		tab.cell(row=row_num, column=13, value=dict[key])
	elif key == 'DATE':
		tab.cell(row=row_num, column=14, value=dict[key])
	elif key == 'ANL_LINK':
		tab.cell(row=row_num, column=15, value=dict[key])
	elif key == 'RUN':
		tab.cell(row=row_num, column=16, value=dict[key])

def write_proj_tab(proj, tab, proj_count):
	tab.cell(row=proj_count+1, column=1, value=proj.title)
	tab.cell(row=proj_count+1, column=2, value=proj.alias)
	tab.cell(row=proj_count+1, column=3, value=proj.description)
	tab.cell(row=proj_count+1, column=4, value=proj.center)
	tab.cell(row=proj_count+1, column=5, value=proj.taxid)
	if proj.opt_att:
		for key in proj.opt_att:
			write_opt_att(proj_count+1, key, proj.opt_att, tab)
def write_analy_tab(analy, tab, analy_count):
	tab.cell(row=analy_count+1, column=1, value=analy.title)
	tab.cell(row=analy_count+1, column=2, value=analy.alias)
	tab.cell(row=analy_count+1, column=3, value=analy.description)
	tab.cell(row=analy_count+1, column=4, value=analy.prj.title)
	tab.cell(row=analy_count+1, column=5, value=analy.experiment)
	tab.cell(row=analy_count+1, column=6, value=analy.ref)
	tab.cell(row=analy_count+1, column=7, value=analy.refmd5)
	if analy.opt_att:
		for key in analy.opt_att:
			write_opt_att(analy_count+1, key, analy.opt_att, tab)

def write_file_tab(file, tab, file_count):
	tab.cell(row=file_count+1, column=1, value=file.analysis.alias)
	tab.cell(row=file_count+1, column=2, value=file.name)
	tab.cell(row=file_count+1, column=3, value=file.filetype)
	tab.cell(row=file_count+1, column=4, value=file.md5)

def write(read_user, user_infos, projects, out):
	wb = openpyxl.load_workbook(filename='src/EVA_Submission_template.V1.1.0.xlsx')
	if read_user:
		ws = wb['Submitter Details']
		ws['A2'] = user_infos['LAST_NAME']
		ws['B2'] = user_infos['FIRST_NAME']
		ws['C2'] = user_infos['TELEPHONE']
		ws['D2'] = user_infos['EMAIL']
		ws['E2'] = user_infos['LABORATORY']
		ws['F2'] = user_infos['CENTER']
		ws['G2'] = user_infos['ADDRESS']
	project_tab = wb['Project']
	analysis_tab = wb['Analysis']
	file_tab = wb['Files']
	proj_count = 1
	analy_count = 1
	file_count = 1
	for proj in projects:
		write_proj_tab(proj, project_tab, proj_count)
		proj_count += 1
		for analy in proj.analyses:
			write_analy_tab(analy, analysis_tab, analy_count)
			analy_count += 1
			for f in analy.files:
				write_file_tab(f, file_tab, file_count)
				file_count += 1
	wb.save(out)
